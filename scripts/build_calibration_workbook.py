"""
build_calibration_workbook.py -- converts
data/eval/human_calibration_disagreement_subset.csv into a formatted Excel
workbook, data/eval/human_calibration_review.xlsx, for Sam to review
directly in Excel.

Read-only conversion: never modifies the source CSV. `cited_chunk_texts`
is stored in the CSV as a JSON-encoded list (a claim can cite more than
one chunk) -- unpacked here into readable, delimited plain text for the
Excel cell, not left as an escaped JSON blob.

Row height is estimated (not truly "auto-fit", which requires rendering
text -- something openpyxl itself cannot do) from character count vs.
column width, so multi-paragraph excerpts are fully visible on open
without the reviewer needing to manually resize every row.

Usage:
    uv run python scripts/build_calibration_workbook.py
"""

import csv
import json
import math
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SOURCE_CSV = PROJECT_ROOT / "data" / "eval" / "human_calibration_disagreement_subset.csv"
OUTPUT_XLSX = PROJECT_ROOT / "data" / "eval" / "human_calibration_review.xlsx"

# Column order per spec, with the judge's own `verdict` column renamed to
# `judge_verdict` for clarity in the workbook (the source CSV's header is
# left untouched -- only this derived copy renames it).
COLUMNS = [
    ("row_number", 8),
    ("judgment_id", 30),
    ("category", 20),
    ("claim_text", 75),
    ("cited_chunk_texts", 75),
    ("judge_verdict", 14),
    ("ai_reviewer_verdict", 16),
    ("human_verdict", 16),
]
SOURCE_TO_HEADER = {"verdict": "judge_verdict"}  # every other name matches as-is

WRAP_COLUMNS = {"claim_text", "cited_chunk_texts"}
LINE_HEIGHT_POINTS = 15  # Excel's default row height for an 11pt font, one line
# Excel's own hard technical ceiling -- 409 points is the real maximum row
# height the format allows, not a stylistic choice made here. Several rows
# in this dataset (multi-chunk claims citing 5-10 full ~1500-char chunks)
# estimate at 1,000-4,000+ points of wrapped text -- no row-height setting
# can make those fully visible without scrolling/clicking into the cell,
# regardless of this script's formatting. The underlying cell VALUE is
# never truncated either way (openpyxl only caps display height, not
# content) -- disclosed plainly in reports.md rather than silently
# under-representing what a capped row can actually show at a glance.
MAX_ROW_HEIGHT = 409


def _format_cited_chunks(raw: str) -> str:
    """Unpacks the JSON-encoded list of cited chunk texts into readable,
    delimited plain text -- one clearly separated block per excerpt,
    instead of a single escaped-JSON string."""
    try:
        chunks = json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return raw
    if not isinstance(chunks, list):
        return raw
    blocks = [
        f"--- excerpt {i} ---\n{chunk}" for i, chunk in enumerate(chunks, start=1)
    ]
    return "\n\n".join(blocks)


def _estimate_lines(text: str, width_chars: int) -> int:
    """Estimates how many wrapped lines `text` will occupy at `width_chars`
    -- accounts for existing newlines (each forces a new line) plus
    character-count wrapping within each line. Not exact (Excel's real
    rendering depends on font metrics, not a fixed char count), but close
    enough to avoid truncated-looking cells on open."""
    if not text:
        return 1
    total = 0
    for line in text.split("\n"):
        total += max(1, math.ceil(len(line) / width_chars))
    return total


def build(source_csv: Path, output_path: Path) -> int:
    with open(source_csv, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    wb = Workbook()
    ws = wb.active
    ws.title = "Calibration Review"

    header_names = [name for name, _ in COLUMNS]
    ws.append(header_names)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    reference_note = (
        "Reference only -- pre-filled from the judge/AI-reviewer pass. "
        "Not for editing; fill human_verdict instead."
    )
    ws["F1"].comment = Comment(reference_note, "build_calibration_workbook.py")
    ws["G1"].comment = Comment(reference_note, "build_calibration_workbook.py")

    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    for row in rows:
        cited = _format_cited_chunks(row["cited_chunk_texts"])
        values = [
            int(row["row_number"]),
            row["judgment_id"],
            row["category"],
            row["claim_text"],
            cited,
            row["verdict"],  # judge_verdict -- blank for the 2 skip rows, same as source
            row["ai_reviewer_verdict"],
            row["human_verdict"],  # blank, same as source
        ]
        ws.append(values)

        excel_row = ws.max_row
        max_lines = 1
        for col_name in WRAP_COLUMNS:
            col_idx = header_names.index(col_name) + 1
            col_width = dict(COLUMNS)[col_name]
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.alignment = wrap_alignment
            max_lines = max(max_lines, _estimate_lines(values[col_idx - 1], col_width))
        ws.row_dimensions[excel_row].height = min(
            MAX_ROW_HEIGHT, max_lines * LINE_HEIGHT_POINTS
        )

    for name, width in COLUMNS:
        col_idx = header_names.index(name) + 1
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    human_verdict_col = header_names.index("human_verdict") + 1
    col_letter = ws.cell(row=1, column=human_verdict_col).column_letter
    dv = DataValidation(
        type="list",
        formula1='"Supported,Partial,Unsupported"',
        allow_blank=True,
        showDropDown=False,  # False = arrow shown (openpyxl's inverted flag naming)
    )
    dv.error = "Choose Supported, Partial, or Unsupported."
    dv.errorTitle = "Invalid entry"
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{ws.max_row}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(rows)


def main() -> None:
    if not SOURCE_CSV.exists():
        print(f"{SOURCE_CSV.relative_to(PROJECT_ROOT)} doesn't exist.", file=sys.stderr)
        sys.exit(1)
    n = build(SOURCE_CSV, OUTPUT_XLSX)
    print(f"[ok] wrote {n} row(s) to {OUTPUT_XLSX.relative_to(PROJECT_ROOT)}")


if __name__ == "__main__":
    main()
